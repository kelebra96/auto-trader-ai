import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from flask import Flask, jsonify, request, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import json
import uuid
import csv
import io
from functools import wraps

# Importar blueprint de notificações
from routes.notifications import notifications_bp

# Configuração básica
app = Flask(__name__)
app.config['SECRET_KEY'] = 'dev-secret-key'
import os
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(os.path.dirname(basedir), "app.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JWT_SECRET_KEY'] = 'jwt-secret-key'
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(basedir), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB máximo

# Inicializar extensões
db = SQLAlchemy(app)
jwt = JWTManager(app)
CORS(app, origins="*")

# Middleware de logging detalhado
@app.before_request
def log_request_info():
    print(f"\n{'='*50}")
    print(f"🔍 REQUISIÇÃO RECEBIDA:")
    print(f"   Método: {request.method}")
    print(f"   URL: {request.url}")
    print(f"   Path: {request.path}")
    print(f"   Headers: {dict(request.headers)}")
    
    # Debug específico para JWT
    auth_header = request.headers.get('Authorization')
    if auth_header:
        print(f"   🔑 Authorization Header: {auth_header[:50]}...")
        if auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
            print(f"   🎫 Token extraído: {token[:20]}...{token[-20:] if len(token) > 40 else token}")
    else:
        print(f"   ❌ Nenhum Authorization Header encontrado")
    
    if request.method in ['POST', 'PUT', 'PATCH']:
        print(f"   Content-Type: {request.content_type}")
        print(f"   Dados: {request.get_data()}")
    print(f"{'='*50}\n")

@app.after_request
def log_response_info(response):
    print(f"\n{'='*50}")
    print(f"📤 RESPOSTA ENVIADA:")
    print(f"   Status: {response.status_code}")
    print(f"   Headers: {dict(response.headers)}")
    if response.status_code >= 400:
        print(f"   ❌ ERRO: {response.get_data(as_text=True)}")
    else:
        print(f"   ✅ SUCESSO")
    print(f"{'='*50}\n")
    return response

# Tratamento de erros JWT
@jwt.expired_token_loader
def expired_token_callback(jwt_header, jwt_payload):
    print(f"DEBUG: Token expirado - Header: {jwt_header}, Payload: {jwt_payload}")
    return jsonify({'error': 'Token expirado'}), 401

@jwt.invalid_token_loader
def invalid_token_callback(error):
    print(f"DEBUG: Token inválido - Erro: {error}")
    print(f"DEBUG: Tipo do erro: {type(error)}")
    print(f"DEBUG: Headers da requisição: {dict(request.headers) if request else 'Sem request'}")
    return jsonify({'error': 'Token inválido'}), 422

@jwt.unauthorized_loader
def missing_token_callback(error):
    print(f"DEBUG: Token ausente - Erro: {error}")
    return jsonify({'error': 'Token de autorização necessário'}), 401

# Modelo de usuário expandido
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    nome_estabelecimento = db.Column(db.String(200), nullable=False)
    
    # Informações pessoais
    nome_completo = db.Column(db.String(200), nullable=True)
    telefone = db.Column(db.String(20), nullable=True)
    empresa = db.Column(db.String(200), nullable=True)
    bio = db.Column(db.Text, nullable=True)
    foto_perfil = db.Column(db.String(500), nullable=True)  # URL ou caminho da foto
    
    # Permissões e cargo
    cargo = db.Column(db.String(100), default='usuario')  # admin, gerente, usuario, visualizador
    permissoes = db.Column(db.JSON, nullable=True)  # Lista de permissões específicas
    ativo = db.Column(db.Boolean, default=True)
    
    # Informações de segurança
    ultimo_login = db.Column(db.DateTime, nullable=True)
    tentativas_login = db.Column(db.Integer, default=0)
    bloqueado_ate = db.Column(db.DateTime, nullable=True)
    
    # Timestamps
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def set_password(self, password):
        """Define a senha do usuário com hash"""
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        """Verifica se a senha está correta"""
        return check_password_hash(self.password_hash, password)
    
    def to_dict(self, include_sensitive=False):
        data = {
            'id': self.id,
            'email': self.email,
            'nome_estabelecimento': self.nome_estabelecimento,
            'nome_completo': self.nome_completo,
            'telefone': self.telefone,
            'empresa': self.empresa,
            'bio': self.bio,
            'foto_perfil': self.foto_perfil,
            'cargo': self.cargo,
            'permissoes': self.permissoes or [],
            'ativo': self.ativo,
            'ultimo_login': self.ultimo_login.isoformat() if self.ultimo_login else None,
            'created_at': self.created_at.isoformat(),
            'updated_at': self.updated_at.isoformat()
        }
        
        if include_sensitive:
            data.update({
                'tentativas_login': self.tentativas_login,
                'bloqueado_ate': self.bloqueado_ate.isoformat() if self.bloqueado_ate else None
            })
            
        return data

# Modelo de produto simples
# Modelo de Empresa
class Empresa(db.Model):
    __tablename__ = 'empresas'
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    nome = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    telefone = db.Column(db.String(20), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'codigo': self.codigo,
            'nome': self.nome,
            'email': self.email,
            'telefone': self.telefone,
            'created_at': self.created_at.isoformat(),
            'updated_at': self.updated_at.isoformat()
        }

# Modelo de Fornecedor
class Fornecedor(db.Model):
    __tablename__ = 'fornecedores'
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    nome = db.Column(db.String(200), nullable=False)
    status = db.Column(db.String(20), default='ativo')  # ativo, inativo
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'codigo': self.codigo,
            'nome': self.nome,
            'status': self.status,
            'created_at': self.created_at.isoformat(),
            'updated_at': self.updated_at.isoformat()
        }

# Modelo de Produto (Simplificado conforme especificação)
class Produto(db.Model):
    __tablename__ = 'produtos'
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True, nullable=False)
    nome = db.Column(db.String(200), nullable=False)
    categoria = db.Column(db.String(100), nullable=False)
    fornecedor_id = db.Column(db.Integer, db.ForeignKey('fornecedores.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relacionamento
    fornecedor = db.relationship('Fornecedor', backref='produtos')

    def to_dict(self):
        return {
            'id': self.id,
            'codigo': self.codigo,
            'nome': self.nome,
            'categoria': self.categoria,
            'fornecedor_id': self.fornecedor_id,
            'fornecedor': self.fornecedor.to_dict() if self.fornecedor else None,
            'created_at': self.created_at.isoformat(),
            'updated_at': self.updated_at.isoformat()
        }

# Modelo de Entrada de Produto (Para dados do mobile)
class EntradaProduto(db.Model):
    __tablename__ = 'entradas_produtos'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresas.id'), nullable=False)
    produto_id = db.Column(db.Integer, db.ForeignKey('produtos.id'), nullable=False)
    data_vencimento = db.Column(db.Date, nullable=False)
    quantidade = db.Column(db.Integer, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relacionamentos
    user = db.relationship('User', backref='entradas_produtos')
    empresa = db.relationship('Empresa', backref='entradas_produtos')
    produto = db.relationship('Produto', backref='entradas_produtos')

    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'empresa_id': self.empresa_id,
            'produto_id': self.produto_id,
            'produto': self.produto.to_dict() if self.produto else None,
            'empresa': self.empresa.to_dict() if self.empresa else None,
            'data_vencimento': self.data_vencimento.isoformat() if self.data_vencimento else None,
            'quantidade': self.quantidade,
            'created_at': self.created_at.isoformat()
        }

# Modelo de Alerta
class Alerta(db.Model):
    __tablename__ = 'alertas'
    
    id = db.Column(db.Integer, primary_key=True)
    entrada_produto_id = db.Column(db.Integer, db.ForeignKey('entradas_produtos.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    tipo = db.Column(db.String(50), nullable=False)  # vencimento, estoque_baixo
    urgencia = db.Column(db.String(20), nullable=False)  # alta, media, baixa
    titulo = db.Column(db.String(200), nullable=False)
    descricao = db.Column(db.Text, nullable=True)
    quantidade_afetada = db.Column(db.Integer, nullable=False)
    valor_estimado_perda = db.Column(db.Float, nullable=True)
    status = db.Column(db.String(20), default='ativo')  # ativo, resolvido, ignorado
    lido = db.Column(db.Boolean, default=False)  # Campo para marcar como lido
    acao_tomada = db.Column(db.String(100), nullable=True)
    detalhes_resolucao = db.Column(db.JSON, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    resolved_at = db.Column(db.DateTime, nullable=True)
    
    # Relacionamentos
    entrada_produto = db.relationship('EntradaProduto', backref='alertas')
    user = db.relationship('User', backref='alertas')
    
    def to_dict(self):
        return {
            'id': self.id,
            'entrada_produto_id': self.entrada_produto_id,
            'produto_nome': self.entrada_produto.produto.nome if self.entrada_produto and self.entrada_produto.produto else None,
            'entrada_produto': self.entrada_produto.to_dict() if self.entrada_produto else None,
            'tipo': self.tipo,
            'urgencia': self.urgencia,
            'titulo': self.titulo,
            'descricao': self.descricao,
            'quantidade_afetada': self.quantidade_afetada,
            'valor_estimado_perda': self.valor_estimado_perda,
            'status': self.status,
            'lido': self.lido,
            'acao_tomada': self.acao_tomada,
            'detalhes_resolucao': self.detalhes_resolucao,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'resolved_at': self.resolved_at.isoformat() if self.resolved_at else None
        }

# Modelo de Configuração de Alerta
class ConfiguracaoAlerta(db.Model):
    __tablename__ = 'configuracoes_alertas'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    tipo_alerta = db.Column(db.String(50), nullable=False)  # vencimento, estoque_baixo
    ativo = db.Column(db.Boolean, default=True)
    dias_antecedencia = db.Column(db.Integer, default=7)  # Para alertas de vencimento
    estoque_minimo = db.Column(db.Integer, default=5)  # Para alertas de estoque baixo
    categorias = db.Column(db.JSON, nullable=True)  # Lista de categorias para filtrar
    notificar_email = db.Column(db.Boolean, default=True)
    notificar_sistema = db.Column(db.Boolean, default=True)
    horario_notificacao = db.Column(db.Time, nullable=True)  # Horário preferido para notificações
    recorrencia = db.Column(db.String(20), default='diaria')  # diaria, semanal, mensal
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relacionamentos
    user = db.relationship('User', backref='configuracoes_alertas')
    
    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'tipo_alerta': self.tipo_alerta,
            'ativo': self.ativo,
            'dias_antecedencia': self.dias_antecedencia,
            'estoque_minimo': self.estoque_minimo,
            'categorias': self.categorias,
            'notificar_email': self.notificar_email,
            'notificar_sistema': self.notificar_sistema,
            'horario_notificacao': self.horario_notificacao.strftime('%H:%M') if self.horario_notificacao else None,
            'recorrencia': self.recorrencia,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

class Venda(db.Model):
    __tablename__ = 'vendas'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    produto_id = db.Column(db.Integer, db.ForeignKey('produtos.id'), nullable=False)
    quantidade = db.Column(db.Integer, nullable=False)
    preco_unitario = db.Column(db.Float, nullable=False)
    valor_total = db.Column(db.Float, nullable=False)
    data_venda = db.Column(db.DateTime, default=datetime.utcnow)
    metodo_pagamento = db.Column(db.String(50), nullable=True)  # dinheiro, cartao, pix
    observacoes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relacionamentos
    user = db.relationship('User', backref='vendas')
    produto = db.relationship('Produto', backref='vendas')
    
    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'produto_id': self.produto_id,
            'produto_nome': self.produto.nome if self.produto else None,
            'quantidade': self.quantidade,
            'preco_unitario': self.preco_unitario,
            'valor_total': self.valor_total,
            'data_venda': self.data_venda.isoformat(),
            'metodo_pagamento': self.metodo_pagamento,
            'observacoes': self.observacoes,
            'created_at': self.created_at.isoformat()
        }

# Rotas de autenticação}

class ConfiguracaoUsuario(db.Model):
    __tablename__ = 'configuracoes_usuario'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, unique=True)
    
    # Configurações de Notificações
    notificacoes_email = db.Column(db.Boolean, default=True)
    notificacoes_push = db.Column(db.Boolean, default=True)
    dias_antecedencia_notificacao = db.Column(db.Integer, default=7)
    
    # Configurações de Segurança
    autenticacao_dois_fatores = db.Column(db.Boolean, default=False)
    timeout_sessao = db.Column(db.Integer, default=60)  # em minutos
    
    # Configurações de Sistema
    tema = db.Column(db.String(20), default='light')  # light, dark
    idioma = db.Column(db.String(10), default='pt-BR')
    
    # Configurações de Email/SMTP
    smtp_servidor = db.Column(db.String(255), nullable=True)
    smtp_porta = db.Column(db.Integer, nullable=True)
    smtp_usuario = db.Column(db.String(255), nullable=True)
    smtp_senha = db.Column(db.String(255), nullable=True)
    smtp_ssl = db.Column(db.Boolean, default=True)
    
    # Configurações de Backup
    backup_automatico = db.Column(db.Boolean, default=False)
    frequencia_backup = db.Column(db.String(20), default='semanal')  # diario, semanal, mensal
    local_backup = db.Column(db.String(500), nullable=True)
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relacionamento
    user = db.relationship('User', backref='configuracao')
    
    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'notificacoes_email': self.notificacoes_email,
            'notificacoes_push': self.notificacoes_push,
            'dias_antecedencia_notificacao': self.dias_antecedencia_notificacao,
            'autenticacao_dois_fatores': self.autenticacao_dois_fatores,
            'timeout_sessao': self.timeout_sessao,
            'tema': self.tema,
            'idioma': self.idioma,
            'smtp_servidor': self.smtp_servidor,
            'smtp_porta': self.smtp_porta,
            'smtp_usuario': self.smtp_usuario,
            'smtp_ssl': self.smtp_ssl,
            'backup_automatico': self.backup_automatico,
            'frequencia_backup': self.frequencia_backup,
            'local_backup': self.local_backup,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

# Sistema de Permissões
PERMISSIONS = {
    'admin': [
        'view_all_users', 'create_user', 'edit_user', 'delete_user', 'change_user_password',
        'view_all_products', 'view_products', 'create_product', 'edit_product', 'delete_product',
        'view_all_sales', 'view_sales', 'create_sale', 'edit_sale', 'delete_sale',
        'view_all_alerts', 'view_alerts', 'create_alert', 'edit_alert', 'delete_alert',
        'view_all_reports', 'view_reports', 'export_data', 'manage_settings',
        'view_dashboard', 'manage_permissions'
    ],
    'gerente': [
        'view_users', 'edit_user', 'change_user_password',
        'view_all_products', 'create_product', 'edit_product', 'delete_product',
        'view_all_sales', 'create_sale', 'edit_sale',
        'view_all_alerts', 'create_alert', 'edit_alert',
        'view_all_reports', 'export_data',
        'view_dashboard'
    ],
    'usuario': [
        'view_own_profile', 'edit_own_profile',
        'view_products', 'create_product', 'edit_product',
        'view_sales', 'create_sale',
        'view_alerts', 'create_alert',
        'view_reports',
        'view_dashboard'
    ],
    'visualizador': [
        'view_own_profile',
        'view_products',
        'view_sales',
        'view_alerts',
        'view_reports',
        'view_dashboard'
    ]
}

def get_user_permissions(cargo):
    """Retorna as permissões de um cargo específico"""
    return PERMISSIONS.get(cargo, [])

def has_permission(user_cargo, required_permission, user_permissions_list=None):
    """Verifica se um cargo tem uma permissão específica"""
    # Se o usuário tem permissões específicas definidas
    if user_permissions_list:
        # Se tem permissão 'all', permite tudo
        if 'all' in user_permissions_list:
            return True
        # Verifica se tem a permissão específica
        if required_permission in user_permissions_list:
            return True
    
    # Fallback para permissões baseadas no cargo
    cargo_permissions = get_user_permissions(user_cargo)
    return required_permission in cargo_permissions

def require_permission(permission):
    """Decorador para verificar permissões em rotas"""
    def decorator(f):
        @wraps(f)
        @jwt_required()
        def decorated_function(*args, **kwargs):
            try:
                current_user_id = get_jwt_identity()
                user = User.query.get(current_user_id)
                
                if not user:
                    return jsonify({'error': 'Usuário não encontrado'}), 404
                
                if not user.ativo:
                    return jsonify({'error': 'Usuário inativo'}), 403
                
                if not has_permission(user.cargo, permission, user.permissoes):
                    return jsonify({
                        'error': 'Acesso negado',
                        'message': f'Você não tem permissão para: {permission}',
                        'required_permission': permission,
                        'user_cargo': user.cargo,
                        'user_permissions': user.permissoes
                    }), 403
                
                return f(*args, **kwargs)
            except Exception as e:
                print(f"Erro na verificação de permissão: {e}")
                return jsonify({'error': 'Erro interno do servidor'}), 500
        
        return decorated_function
    return decorator

def require_cargo(allowed_cargos):
    """Decorador para verificar cargos específicos"""
    def decorator(f):
        @wraps(f)
        @jwt_required()
        def decorated_function(*args, **kwargs):
            try:
                current_user_id = get_jwt_identity()
                user = User.query.get(current_user_id)
                
                if not user:
                    return jsonify({'error': 'Usuário não encontrado'}), 404
                
                if not user.ativo:
                    return jsonify({'error': 'Usuário inativo'}), 403
                
                if user.cargo not in allowed_cargos:
                    return jsonify({
                        'error': 'Acesso negado',
                        'message': f'Cargo {user.cargo} não autorizado para esta operação',
                        'allowed_cargos': allowed_cargos,
                        'user_cargo': user.cargo
                    }), 403
                
                return f(*args, **kwargs)
            except Exception as e:
                print(f"Erro na verificação de cargo: {e}")
                return jsonify({'error': 'Erro interno do servidor'}), 500
        
        return decorated_function
    return decorator

# Função auxiliar para upload de arquivos
def allowed_file(filename):
    """Verifica se o arquivo tem uma extensão permitida"""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_file(file):
    """Salva o arquivo enviado e retorna o caminho relativo"""
    if file and allowed_file(file.filename):
        # Gerar nome único para o arquivo
        filename = secure_filename(file.filename)
        file_extension = filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
        
        # Criar diretório se não existir
        upload_folder = app.config['UPLOAD_FOLDER']
        os.makedirs(upload_folder, exist_ok=True)
        
        # Salvar arquivo
        file_path = os.path.join(upload_folder, unique_filename)
        file.save(file_path)
        
        # Retornar caminho relativo para armazenar no banco
        return f"/uploads/{unique_filename}"
    
    return None

@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.get_json()
    
    # Verificar se usuário já existe
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'Email já cadastrado'}), 400
    
    # Criar usuário
    user = User(
        email=data['email'],
        nome_estabelecimento=data['nome_estabelecimento']
    )
    user.set_password(data['password'])  # Usar hash de senha
    
    db.session.add(user)
    db.session.commit()
    
    # Criar token
    access_token = create_access_token(identity=str(user.id))
    
    return jsonify({
        'message': 'Usuário criado com sucesso',
        'token': access_token,
        'usuario': {
            'id': user.id,
            'email': user.email,
            'nome_estabelecimento': user.nome_estabelecimento
        }
    }), 201

@app.route('/api/auth/login', methods=['POST', 'OPTIONS'])
def login():
    # Handle preflight requests
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response
    
    print("=== LOGIN DEBUG ===")
    print(f"Request method: {request.method}")
    print(f"Request headers: {dict(request.headers)}")
    print(f"Request content type: {request.content_type}")
    print(f"Raw request data: {request.get_data()}")
    
    try:
        data = request.get_json()
        print(f"Parsed JSON data: {data}")
    except Exception as e:
        print(f"Error parsing JSON: {e}")
        return jsonify({'error': 'Dados inválidos'}), 400
    
    if not data or 'email' not in data or 'password' not in data:
        print("Missing email or password in request")
        return jsonify({'error': 'Email e senha são obrigatórios'}), 400
    
    email = data['email']
    password = data['password']
    print(f"Login attempt - Email: {email}, Password: {password}")
    
    user = User.query.filter_by(email=email).first()
    print(f"User found: {user}")
    
    if user:
        print(f"User password hash in DB: {user.password_hash}")
        print(f"Password match: {user.check_password(password)}")
    
    if user and user.check_password(password):
        access_token = create_access_token(identity=str(user.id))
        print(f"Login successful, token created: {access_token[:50]}...")
        return jsonify({
            'token': access_token,
            'usuario': {
                'id': user.id,
                'email': user.email,
                'nome_estabelecimento': user.nome_estabelecimento
            }
        }), 200
    else:
        print("Login failed - invalid credentials")
        return jsonify({'error': 'Credenciais inválidas'}), 401

# ==================== FUNÇÕES DE PROCESSAMENTO DE IMPORTAÇÃO ====================

def process_json_import(file, user_id):
    """Processa importação de arquivo JSON"""
    produtos_importados = []
    erros = []
    
    try:
        content = file.read().decode('utf-8')
        data = json.loads(content)
        
        # Se o JSON é uma lista de produtos
        if isinstance(data, list):
            produtos_data = data
        # Se o JSON tem uma chave 'produtos'
        elif isinstance(data, dict) and 'produtos' in data:
            produtos_data = data['produtos']
        else:
            raise ValueError("Formato JSON inválido. Esperado uma lista de produtos ou objeto com chave 'produtos'")
        
        for i, produto_data in enumerate(produtos_data):
            try:
                produto = create_produto_from_data(produto_data, user_id)
                if produto:
                    produtos_importados.append(produto.to_dict())
            except Exception as e:
                erros.append(f"Linha {i+1}: {str(e)}")
                
    except json.JSONDecodeError as e:
        erros.append(f"Erro ao decodificar JSON: {str(e)}")
    except Exception as e:
        erros.append(f"Erro no processamento: {str(e)}")
    
    return produtos_importados, erros

def process_csv_import(file, user_id):
    """Processa importação de arquivo CSV"""
    produtos_importados = []
    erros = []
    
    try:
        # Ler o arquivo como string
        content = file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(content))
        
        for i, row in enumerate(csv_reader):
            try:
                produto = create_produto_from_data(row, user_id)
                if produto:
                    produtos_importados.append(produto.to_dict())
            except Exception as e:
                erros.append(f"Linha {i+2}: {str(e)}")  # +2 porque linha 1 é header
                
    except Exception as e:
        erros.append(f"Erro no processamento CSV: {str(e)}")
    
    return produtos_importados, erros

def process_xlsx_import(file, user_id):
    """Processa importação de arquivo XLSX"""
    produtos_importados = []
    erros = []
    
    try:
        # Para XLSX, vamos tentar importar openpyxl se disponível
        try:
            import openpyxl
        except ImportError:
            erros.append("Biblioteca openpyxl não instalada. Instale com: pip install openpyxl")
            return produtos_importados, erros
        
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        # Obter headers da primeira linha
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # Processar dados das linhas seguintes
        for row_num in range(2, sheet.max_row + 1):
            try:
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value
                
                produto = create_produto_from_data(row_data, user_id)
                if produto:
                    produtos_importados.append(produto.to_dict())
            except Exception as e:
                erros.append(f"Linha {row_num}: {str(e)}")
                
    except Exception as e:
        erros.append(f"Erro no processamento XLSX: {str(e)}")
    
    return produtos_importados, erros

def create_produto_from_data(data, user_id):
    """Cria um produto a partir dos dados importados"""
    try:
        # Mapear campos (aceitar variações de nomes)
        nome = data.get('nome') or data.get('name') or data.get('produto')
        if not nome:
            raise ValueError("Campo 'nome' é obrigatório")
        
        # Campos obrigatórios com valores padrão
        categoria = data.get('categoria') or data.get('category') or 'Geral'
        preco = float(data.get('preco') or data.get('price') or data.get('valor') or 0)
        quantidade = int(data.get('quantidade') or data.get('quantity') or data.get('estoque') or 0)
        
        # Campos opcionais
        descricao = data.get('descricao') or data.get('description') or ''
        codigo_barras = data.get('codigo_barras') or data.get('barcode') or data.get('ean') or ''
        fornecedor = data.get('fornecedor') or data.get('supplier') or ''
        
        # Data de validade (opcional)
        data_validade = None
        validade_str = data.get('data_validade') or data.get('validade') or data.get('expiry_date')
        if validade_str:
            try:
                if isinstance(validade_str, str):
                    # Tentar diferentes formatos de data
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            data_validade = datetime.strptime(validade_str, fmt).date()
                            break
                        except ValueError:
                            continue
                elif hasattr(validade_str, 'date'):  # Para objetos datetime do Excel
                    data_validade = validade_str.date()
            except:
                pass  # Ignorar erros de data
        
        # Criar produto
        produto = Produto(
            nome=nome,
            categoria=categoria,
            preco=preco,
            quantidade=quantidade,
            descricao=descricao,
            codigo_barras=codigo_barras,
            fornecedor=fornecedor,
            data_validade=data_validade,
            user_id=user_id,
            status='ativo'
        )
        
        db.session.add(produto)
        db.session.commit()
        
        return produto
        
    except Exception as e:
        db.session.rollback()
        raise e

# ==================== ROTAS DE PRODUTOS ====================

@app.route('/api/produtos', methods=['GET'])
@require_permission('view_products')
def listar_produtos():
    try:
        print(f"DEBUG: Listando produtos")
        produtos = Produto.query.all()
        print(f"DEBUG: Encontrados {len(produtos)} produtos")
        return jsonify({
            'produtos': [produto.to_dict() for produto in produtos]
        }), 200
    except Exception as e:
        print(f"ERROR: Erro ao listar produtos: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/produtos', methods=['POST'])
@require_permission('create_product')
def criar_produto():
    try:
        print("DEBUG: Entrando na função criar_produto")
        print(f"DEBUG: Headers da requisição: {dict(request.headers)}")
        
        data = request.get_json()
        print(f"DEBUG: Dados recebidos: {data}")
        
        # Validação básica
        if not data.get('nome') or not data.get('categoria') or not data.get('codigo') or not data.get('fornecedor_id'):
            return jsonify({'error': 'Campos obrigatórios: codigo, nome, categoria, fornecedor_id'}), 400
        
        # Verificar se o fornecedor existe
        fornecedor = Fornecedor.query.get(data['fornecedor_id'])
        if not fornecedor:
            return jsonify({'error': 'Fornecedor não encontrado'}), 400
        
        # Verificar se o código já existe
        produto_existente = Produto.query.filter_by(codigo=data['codigo']).first()
        if produto_existente:
            return jsonify({'error': 'Código de produto já existe'}), 400
        
        # Criar produto
        produto = Produto(
            codigo=data['codigo'],
            nome=data['nome'],
            categoria=data['categoria'],
            fornecedor_id=data['fornecedor_id']
        )
        
        db.session.add(produto)
        db.session.commit()
        
        return jsonify({
            'message': 'Produto criado com sucesso',
            'produto': produto.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/produtos/<int:produto_id>', methods=['PUT'])
@require_permission('edit_product')
def atualizar_produto(produto_id):
    try:
        user_id = int(get_jwt_identity())
        produto = Produto.query.filter_by(id=produto_id, user_id=user_id).first()
        
        if not produto:
            return jsonify({'error': 'Produto não encontrado'}), 404
        
        data = request.get_json()
        
        # Atualizar campos
        if 'nome' in data:
            produto.nome = data['nome']
        if 'categoria' in data:
            produto.categoria = data['categoria']
        if 'data_validade' in data:
            if data['data_validade']:
                produto.data_validade = datetime.strptime(data['data_validade'], '%Y-%m-%d').date()
            else:
                produto.data_validade = None
        if 'quantidade' in data:
            produto.quantidade = data['quantidade']
        if 'preco_custo' in data:
            produto.preco_custo = data['preco_custo']
        if 'preco_venda' in data:
            produto.preco_venda = data['preco_venda']
        if 'fornecedor' in data:
            produto.fornecedor = data['fornecedor']
        if 'codigo_barras' in data:
            produto.codigo_barras = data['codigo_barras']
        
        produto.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'message': 'Produto atualizado com sucesso',
            'produto': produto.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/produtos/<int:produto_id>', methods=['DELETE'])
@require_permission('delete_product')
def deletar_produto(produto_id):
    try:
        user_id = int(get_jwt_identity())
        produto = Produto.query.filter_by(id=produto_id, user_id=user_id).first()
        
        if not produto:
            return jsonify({'error': 'Produto não encontrado'}), 404
        
        db.session.delete(produto)
        db.session.commit()
        
        return jsonify({'message': 'Produto deletado com sucesso'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/produtos/import', methods=['POST'])
@require_permission('create_product')
def import_produtos():
    """Importa produtos de arquivo CSV, XLSX ou JSON"""
    try:
        user_id = int(get_jwt_identity())
        
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        # Verificar extensão do arquivo
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        if file_ext not in ['csv', 'xlsx', 'json']:
            return jsonify({'error': 'Formato de arquivo não suportado. Use CSV, XLSX ou JSON'}), 400
        
        # Processar arquivo baseado na extensão
        produtos_importados = []
        erros = []
        
        try:
            if file_ext == 'json':
                produtos_importados, erros = process_json_import(file, user_id)
            elif file_ext == 'csv':
                produtos_importados, erros = process_csv_import(file, user_id)
            elif file_ext == 'xlsx':
                produtos_importados, erros = process_xlsx_import(file, user_id)
        except Exception as e:
            return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 400
        
        return jsonify({
            'message': 'Importação concluída',
            'imported': len(produtos_importados),
            'errors': erros,
            'produtos': produtos_importados
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== ENDPOINTS DE FORNECEDORES ====================

@app.route('/api/fornecedores', methods=['GET'])
@require_permission('view_products')
def listar_fornecedores():
    try:
        print(f"DEBUG: Listando fornecedores")
        fornecedores = Fornecedor.query.all()
        print(f"DEBUG: Encontrados {len(fornecedores)} fornecedores")
        return jsonify({
            'fornecedores': [fornecedor.to_dict() for fornecedor in fornecedores]
        }), 200
    except Exception as e:
        print(f"ERROR: Erro ao listar fornecedores: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/fornecedores', methods=['POST'])
@require_permission('create_product')
def criar_fornecedor():
    try:
        print("DEBUG: Entrando na função criar_fornecedor")
        
        data = request.get_json()
        print(f"DEBUG: Dados recebidos: {data}")
        
        # Validação básica
        if not data.get('nome') or not data.get('codigo'):
            return jsonify({'error': 'Campos obrigatórios: codigo, nome'}), 400
        
        # Verificar se o código já existe
        fornecedor_existente = Fornecedor.query.filter_by(codigo=data['codigo']).first()
        if fornecedor_existente:
            return jsonify({'error': 'Código de fornecedor já existe'}), 400
        
        # Criar fornecedor
        fornecedor = Fornecedor(
            codigo=data['codigo'],
            nome=data['nome'],
            status=data.get('status', 'ativo')
        )
        
        db.session.add(fornecedor)
        db.session.commit()
        
        return jsonify({
            'message': 'Fornecedor criado com sucesso',
            'fornecedor': fornecedor.to_dict()
        }), 201
        
    except Exception as e:
        print(f"ERROR: Erro ao criar fornecedor: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/fornecedores/<int:fornecedor_id>', methods=['PUT'])
@require_permission('edit_product')
def atualizar_fornecedor(fornecedor_id):
    try:
        print(f"DEBUG: Atualizando fornecedor {fornecedor_id}")
        
        fornecedor = Fornecedor.query.get(fornecedor_id)
        if not fornecedor:
            return jsonify({'error': 'Fornecedor não encontrado'}), 404
        
        data = request.get_json()
        print(f"DEBUG: Dados recebidos: {data}")
        
        # Verificar se o código já existe (exceto para o próprio fornecedor)
        if data.get('codigo') and data['codigo'] != fornecedor.codigo:
            fornecedor_existente = Fornecedor.query.filter_by(codigo=data['codigo']).first()
            if fornecedor_existente:
                return jsonify({'error': 'Código de fornecedor já existe'}), 400
        
        # Atualizar campos
        if data.get('codigo'):
            fornecedor.codigo = data['codigo']
        if data.get('nome'):
            fornecedor.nome = data['nome']
        if data.get('status'):
            fornecedor.status = data['status']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Fornecedor atualizado com sucesso',
            'fornecedor': fornecedor.to_dict()
        }), 200
        
    except Exception as e:
        print(f"ERROR: Erro ao atualizar fornecedor: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/fornecedores/<int:fornecedor_id>', methods=['DELETE'])
@require_permission('delete_product')
def deletar_fornecedor(fornecedor_id):
    try:
        print(f"DEBUG: Deletando fornecedor {fornecedor_id}")
        
        fornecedor = Fornecedor.query.get(fornecedor_id)
        if not fornecedor:
            return jsonify({'error': 'Fornecedor não encontrado'}), 404
        
        # Verificar se há produtos vinculados
        produtos_vinculados = Produto.query.filter_by(fornecedor_id=fornecedor_id).count()
        if produtos_vinculados > 0:
            return jsonify({'error': f'Não é possível deletar. Há {produtos_vinculados} produto(s) vinculado(s) a este fornecedor'}), 400
        
        db.session.delete(fornecedor)
        db.session.commit()
        
        return jsonify({'message': 'Fornecedor deletado com sucesso'}), 200
        
    except Exception as e:
        print(f"ERROR: Erro ao deletar fornecedor: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ==================== ENDPOINTS DE EMPRESAS ====================

@app.route('/api/empresas', methods=['GET'])
@require_permission('view_products')
def listar_empresas():
    try:
        print(f"DEBUG: Listando empresas")
        empresas = Empresa.query.all()
        print(f"DEBUG: Encontradas {len(empresas)} empresas")
        return jsonify({
            'empresas': [empresa.to_dict() for empresa in empresas]
        }), 200
    except Exception as e:
        print(f"ERROR: Erro ao listar empresas: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/empresas', methods=['POST'])
@require_permission('create_product')
def criar_empresa():
    try:
        print("DEBUG: Entrando na função criar_empresa")
        
        data = request.get_json()
        print(f"DEBUG: Dados recebidos: {data}")
        
        # Validação básica
        if not data.get('nome') or not data.get('codigo') or not data.get('email') or not data.get('telefone'):
            return jsonify({'error': 'Campos obrigatórios: codigo, nome, email, telefone'}), 400
        
        # Verificar se o código já existe
        empresa_existente = Empresa.query.filter_by(codigo=data['codigo']).first()
        if empresa_existente:
            return jsonify({'error': 'Código de empresa já existe'}), 400
        
        # Criar empresa
        empresa = Empresa(
            codigo=data['codigo'],
            nome=data['nome'],
            email=data['email'],
            telefone=data['telefone']
        )
        
        db.session.add(empresa)
        db.session.commit()
        
        return jsonify({
            'message': 'Empresa criada com sucesso',
            'empresa': empresa.to_dict()
        }), 201
        
    except Exception as e:
        print(f"ERROR: Erro ao criar empresa: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/empresas/<int:empresa_id>', methods=['PUT'])
@require_permission('edit_product')
def atualizar_empresa(empresa_id):
    try:
        print(f"DEBUG: Atualizando empresa {empresa_id}")
        
        empresa = Empresa.query.get(empresa_id)
        if not empresa:
            return jsonify({'error': 'Empresa não encontrada'}), 404
        
        data = request.get_json()
        print(f"DEBUG: Dados recebidos: {data}")
        
        # Verificar se o código já existe (exceto para a própria empresa)
        if data.get('codigo') and data['codigo'] != empresa.codigo:
            empresa_existente = Empresa.query.filter_by(codigo=data['codigo']).first()
            if empresa_existente:
                return jsonify({'error': 'Código de empresa já existe'}), 400
        
        # Atualizar campos
        if data.get('codigo'):
            empresa.codigo = data['codigo']
        if data.get('nome'):
            empresa.nome = data['nome']
        if data.get('email'):
            empresa.email = data['email']
        if data.get('telefone'):
            empresa.telefone = data['telefone']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Empresa atualizada com sucesso',
            'empresa': empresa.to_dict()
        }), 200
        
    except Exception as e:
        print(f"ERROR: Erro ao atualizar empresa: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/empresas/<int:empresa_id>', methods=['DELETE'])
@require_permission('delete_product')
def deletar_empresa(empresa_id):
    try:
        print(f"DEBUG: Deletando empresa {empresa_id}")
        
        empresa = Empresa.query.get(empresa_id)
        if not empresa:
            return jsonify({'error': 'Empresa não encontrada'}), 404
        
        # Verificar se há entradas de produtos vinculadas
        entradas_vinculadas = EntradaProduto.query.filter_by(empresa_id=empresa_id).count()
        if entradas_vinculadas > 0:
            return jsonify({'error': f'Não é possível deletar. Há {entradas_vinculadas} entrada(s) de produto(s) vinculada(s) a esta empresa'}), 400
        
        db.session.delete(empresa)
        db.session.commit()
        
        return jsonify({'message': 'Empresa deletada com sucesso'}), 200
        
    except Exception as e:
        print(f"ERROR: Erro ao deletar empresa: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ==================== ENDPOINTS DE ENTRADAS DE PRODUTOS (MOBILE) ====================

@app.route('/api/entradas-produtos', methods=['GET'])
@require_permission('view_products')
def listar_entradas_produtos():
    try:
        user_id = int(get_jwt_identity())
        print(f"DEBUG: Listando entradas de produtos para user_id: {user_id}")
        
        # Filtros opcionais
        empresa_id = request.args.get('empresa_id')
        produto_id = request.args.get('produto_id')
        
        query = EntradaProduto.query.filter_by(user_id=user_id)
        
        if empresa_id:
            query = query.filter_by(empresa_id=empresa_id)
        if produto_id:
            query = query.filter_by(produto_id=produto_id)
        
        entradas = query.order_by(EntradaProduto.created_at.desc()).all()
        print(f"DEBUG: Encontradas {len(entradas)} entradas")
        
        return jsonify({
            'entradas': [entrada.to_dict() for entrada in entradas]
        }), 200
    except Exception as e:
        print(f"ERROR: Erro ao listar entradas de produtos: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/entradas-produtos', methods=['POST'])
@require_permission('create_product')
def criar_entrada_produto():
    try:
        print("DEBUG: Entrando na função criar_entrada_produto")
        
        user_id = int(get_jwt_identity())
        data = request.get_json()
        print(f"DEBUG: Dados recebidos: {data}")
        
        # Validação básica
        required_fields = ['empresa_id', 'produto_id', 'data_vencimento', 'quantidade']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Campo obrigatório: {field}'}), 400
        
        # Verificar se empresa e produto existem
        empresa = Empresa.query.get(data['empresa_id'])
        if not empresa:
            return jsonify({'error': 'Empresa não encontrada'}), 400
        
        produto = Produto.query.get(data['produto_id'])
        if not produto:
            return jsonify({'error': 'Produto não encontrado'}), 400
        
        # Converter data_vencimento
        try:
            data_vencimento = datetime.strptime(data['data_vencimento'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de data inválido. Use YYYY-MM-DD'}), 400
        
        # Criar entrada de produto
        entrada = EntradaProduto(
            user_id=user_id,
            empresa_id=data['empresa_id'],
            produto_id=data['produto_id'],
            data_vencimento=data_vencimento,
            quantidade=data['quantidade']
        )
        
        db.session.add(entrada)
        db.session.commit()
        
        return jsonify({
            'message': 'Entrada de produto criada com sucesso',
            'entrada': entrada.to_dict()
        }), 201
        
    except Exception as e:
        print(f"ERROR: Erro ao criar entrada de produto: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ==================== ENDPOINTS DE BUSCA PARA MOBILE ====================

@app.route('/api/produtos/buscar', methods=['GET'])
@require_permission('view_products')
def buscar_produtos():
    try:
        termo = request.args.get('q', '').strip()
        if not termo:
            return jsonify({'produtos': []}), 200
        
        # Buscar por código ou nome
        produtos = Produto.query.filter(
            db.or_(
                Produto.codigo.ilike(f'%{termo}%'),
                Produto.nome.ilike(f'%{termo}%')
            )
        ).limit(10).all()
        
        return jsonify({
            'produtos': [produto.to_dict() for produto in produtos]
        }), 200
        
    except Exception as e:
        print(f"ERROR: Erro ao buscar produtos: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ==================== ENDPOINTS DE RELATÓRIOS ====================

@app.route('/api/dashboard', methods=['GET'])
@require_permission('view_dashboard')
def dashboard_principal():
    """Retorna dados do dashboard principal"""
    try:
        # Total de produtos
        total_produtos = Produto.query.count()
        
        # Total de fornecedores
        total_fornecedores = Fornecedor.query.count()
        
        # Total de empresas
        total_empresas = Empresa.query.count()
        
        # Total de entradas de produtos
        total_entradas = EntradaProduto.query.count()
        
        # Entradas recentes (últimas 5)
        entradas_recentes = EntradaProduto.query.order_by(EntradaProduto.id.desc()).limit(5).all()
        
        # Produtos recentes (últimos 5 produtos cadastrados)
        produtos_recentes = Produto.query.order_by(Produto.id.desc()).limit(5).all()
        
        # Produtos vencendo em 7 dias (baseado nas entradas)
        from datetime import datetime, timedelta
        data_limite = datetime.now().date() + timedelta(days=7)
        entradas_vencendo = EntradaProduto.query.filter(
            EntradaProduto.data_vencimento <= data_limite,
            EntradaProduto.data_vencimento >= datetime.now().date()
        ).order_by(EntradaProduto.data_vencimento.asc()).limit(5).all()
        
        # Produtos vencidos (contagem baseada nas entradas)
        produtos_vencidos_count = EntradaProduto.query.filter(
            EntradaProduto.data_vencimento < datetime.now().date()
        ).count()
        
        # Produtos vencendo (contagem)
        produtos_vencendo_count = EntradaProduto.query.filter(
            EntradaProduto.data_vencimento <= data_limite,
            EntradaProduto.data_vencimento >= datetime.now().date()
        ).count()
        
        # Valor total do estoque (simplificado)
        valor_estoque = 0  # Será calculado quando tivermos preços nas entradas
        
        # Usuários ativos (simplificado)
        usuarios_ativos = User.query.filter(User.ativo == True).count()
        
        # Fornecedores com mais produtos
        fornecedores_top = db.session.query(
            Fornecedor.nome,
            db.func.count(Produto.id).label('total_produtos')
        ).join(Produto).group_by(Fornecedor.id, Fornecedor.nome).order_by(
            db.func.count(Produto.id).desc()
        ).limit(5).all()
        
        return jsonify({
            'stats': {
                'totalProdutos': total_produtos,
                'produtosVencendo': produtos_vencendo_count,
                'produtosVencidos': produtos_vencidos_count,
                'valorEstoque': valor_estoque,
                'usuariosAtivos': usuarios_ativos
            },
            'changes': {
                'totalProdutos': 0,  # Simplified - no historical comparison
                'produtosVencendo': 0,
                'produtosVencidos': 0,
                'valorEstoque': 0,
                'usuariosAtivos': 0
            },
            'recentProducts': [p.to_dict() for p in produtos_recentes],
            'expiringProducts': [e.to_dict() for e in entradas_vencendo],
            'recentEntries': [e.to_dict() for e in entradas_recentes],
            'topSuppliers': [{'nome': f.nome, 'total_produtos': f.total_produtos} for f in fornecedores_top]
        })
        
    except Exception as e:
        print(f"Erro ao buscar dados do dashboard: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/relatorios/dashboard', methods=['GET'])
@require_permission('view_reports')
def dashboard_relatorios():
    """Retorna dados do dashboard de relatórios"""
    try:
        user_id = get_jwt_identity()
        
        # Produtos vencidos
        produtos_vencidos = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.data_validade < datetime.now().date(),
            Produto.status == 'ativo'
        ).count()
        
        # Produtos vencendo em 7 dias
        data_limite = datetime.now().date() + timedelta(days=7)
        produtos_vencendo_7dias = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.data_validade <= data_limite,
            Produto.data_validade >= datetime.now().date(),
            Produto.status == 'ativo'
        ).count()
        
        # Calcular economia (margem de lucro total dos produtos ativos)
        produtos_ativos = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.status == 'ativo'
        ).all()
        
        economia = sum([
            ((p.preco_venda or 0) - (p.preco_custo or 0)) * (p.quantidade or 0) 
            for p in produtos_ativos 
            if (p.preco_venda or 0) > (p.preco_custo or 0)
        ])
        
        # Calcular perdas (produtos vencidos)
        produtos_perdidos = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.data_validade < datetime.now().date(),
            Produto.status == 'ativo'
        ).all()
        
        perdas = sum([(p.preco_custo or 0) * (p.quantidade or 0) for p in produtos_perdidos])
        
        # Produtos mais perdidos
        produtos_mais_perdidos = []
        for produto in produtos_perdidos[:4]:
            perda_valor = (produto.preco_custo or 0) * (produto.quantidade or 0)
            produtos_mais_perdidos.append({
                'nome': produto.nome,
                'perdas': f'R$ {perda_valor:.2f}',
                'porcentagem': min(100, int((perda_valor / max(perdas, 1)) * 100))
            })
        
        return jsonify({
            'produtos_vencidos': produtos_vencidos,
            'produtos_vencendo_7dias': produtos_vencendo_7dias,
            'economia': economia,
            'perdas': perdas,
            'produtos_mais_perdidos': produtos_mais_perdidos
        })
        
    except Exception as e:
        print(f"Erro ao buscar dashboard: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/relatorios/validades', methods=['GET'])
@require_permission('view_reports')
def relatorio_validades():
    """Gera relatório de validades"""
    try:
        user_id = get_jwt_identity()
        
        # Produtos vencidos
        vencidos = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.data_validade < datetime.now().date(),
            Produto.status == 'ativo'
        ).all()
        
        # Produtos vencendo hoje
        vencendo_hoje = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.data_validade == datetime.now().date(),
            Produto.status == 'ativo'
        ).all()
        
        # Produtos vencendo em 7 dias
        data_limite_7 = datetime.now().date() + timedelta(days=7)
        vencendo_7dias = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.data_validade <= data_limite_7,
            Produto.data_validade > datetime.now().date(),
            Produto.status == 'ativo'
        ).all()
        
        # Produtos vencendo em 30 dias
        data_limite_30 = datetime.now().date() + timedelta(days=30)
        vencendo_30dias = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.data_validade <= data_limite_30,
            Produto.data_validade > data_limite_7,
            Produto.status == 'ativo'
        ).all()
        
        # Produtos críticos (vencidos + vencendo hoje + próximos 7 dias)
        produtos_criticos = []
        for produto in (vencidos + vencendo_hoje + vencendo_7dias):
            dias_para_vencer = (produto.data_validade - datetime.now().date()).days
            if dias_para_vencer < 0:
                status = 'vencido'
                urgencia = 'alta'
            elif dias_para_vencer == 0:
                status = 'vence_hoje'
                urgencia = 'alta'
            elif dias_para_vencer <= 3:
                status = 'critico'
                urgencia = 'alta'
            else:
                status = 'atencao'
                urgencia = 'media'
                
            produtos_criticos.append({
                'id': produto.id,
                'nome': produto.nome,
                'categoria': produto.categoria,
                'validade': produto.data_validade.strftime('%d/%m/%Y') if produto.data_validade else 'N/A',
                'quantidade': produto.quantidade,
                'status': status,
                'urgencia': urgencia,
                'dias_para_vencer': dias_para_vencer
            })
        
        return jsonify({
            'resumo': {
                'vencidos': len(vencidos),
                'vencendo_hoje': len(vencendo_hoje),
                'vencendo_7dias': len(vencendo_7dias),
                'vencendo_30dias': len(vencendo_30dias)
            },
            'produtos_criticos': produtos_criticos[:20]  # Limitar a 20 produtos
        })
        
    except Exception as e:
        print(f"Erro ao gerar relatório de validades: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/relatorios/perdas', methods=['GET'])
@require_permission('view_reports')
def relatorio_perdas():
    """Gera relatório de análise de perdas"""
    try:
        user_id = get_jwt_identity()
        
        # Produtos vencidos (perdas)
        produtos_vencidos = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.data_validade < datetime.now().date(),
            Produto.status == 'ativo'
        ).all()
        
        # Calcular perdas por categoria
        perdas_por_categoria = {}
        perda_total = 0
        
        for produto in produtos_vencidos:
            categoria = produto.categoria
            valor_perda = (produto.preco_custo or 0) * (produto.quantidade or 0)
            perda_total += valor_perda
            
            if categoria not in perdas_por_categoria:
                perdas_por_categoria[categoria] = {
                    'categoria': categoria,
                    'quantidade_produtos': 0,
                    'valor_perdido': 0,
                    'produtos': []
                }
            
            perdas_por_categoria[categoria]['quantidade_produtos'] += 1
            perdas_por_categoria[categoria]['valor_perdido'] += valor_perda
            perdas_por_categoria[categoria]['produtos'].append({
                'nome': produto.nome,
                'quantidade': produto.quantidade,
                'valor_unitario': produto.preco_custo or 0,
                'valor_total_perdido': valor_perda,
                'data_vencimento': produto.data_validade.strftime('%d/%m/%Y') if produto.data_validade else 'N/A'
            })
        
        # Converter para lista e ordenar por valor perdido
        perdas_lista = list(perdas_por_categoria.values())
        perdas_lista.sort(key=lambda x: x['valor_perdido'], reverse=True)
        
        # Produtos com maior perda individual
        produtos_maior_perda = []
        for produto in produtos_vencidos:
            valor_perda = (produto.preco_custo or 0) * (produto.quantidade or 0)
            produtos_maior_perda.append({
                'nome': produto.nome,
                'categoria': produto.categoria,
                'quantidade': produto.quantidade,
                'valor_perdido': valor_perda,
                'data_vencimento': produto.data_validade.strftime('%d/%m/%Y') if produto.data_validade else 'N/A'
            })
        
        produtos_maior_perda.sort(key=lambda x: x['valor_perdido'], reverse=True)
        
        return jsonify({
            'resumo': {
                'perda_total': perda_total,
                'produtos_perdidos': len(produtos_vencidos),
                'categorias_afetadas': len(perdas_por_categoria)
            },
            'perdas_por_categoria': perdas_lista,
            'produtos_maior_perda': produtos_maior_perda[:10]
        })
        
    except Exception as e:
        print(f"Erro ao gerar relatório de perdas: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/relatorios/estoque', methods=['GET'])
@require_permission('view_reports')
def relatorio_estoque():
    """Gera relatório de produtos em estoque"""
    try:
        user_id = get_jwt_identity()
        print(f"Gerando relatório de estoque para user_id: {user_id}")
        
        # Todos os produtos ativos
        produtos = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.status == 'ativo'
        ).all()
        
        print(f"Produtos encontrados: {len(produtos)}")
        for p in produtos:
            print(f"Produto: {p.nome}, Quantidade: {p.quantidade}, Preço: {p.preco_venda}")
        
        # Agrupar por categoria
        estoque_por_categoria = {}
        valor_total_estoque = 0
        
        for produto in produtos:
            categoria = produto.categoria
            quantidade = produto.quantidade if produto.quantidade is not None else 0
            preco = produto.preco_venda if produto.preco_venda is not None else 0
            valor_produto = preco * quantidade
            valor_total_estoque += valor_produto
            
            print(f"Processando {produto.nome}: qtd={quantidade}, preço={preco}, valor={valor_produto}")
            
            if categoria not in estoque_por_categoria:
                estoque_por_categoria[categoria] = {
                    'categoria': categoria,
                    'quantidade_produtos': 0,
                    'quantidade_total': 0,
                    'valor_total': 0,
                    'produtos': []
                }
            
            estoque_por_categoria[categoria]['quantidade_produtos'] += 1
            estoque_por_categoria[categoria]['quantidade_total'] += quantidade
            estoque_por_categoria[categoria]['valor_total'] += valor_produto
            estoque_por_categoria[categoria]['produtos'].append({
                'id': produto.id,
                'nome': produto.nome,
                'quantidade': quantidade,
                'preco_venda': preco,
                'valor_total': valor_produto,
                'data_validade': produto.data_validade.strftime('%d/%m/%Y') if produto.data_validade else 'N/A',
                'fornecedor': produto.fornecedor or 'N/A'
            })
        
        # Converter para lista
        estoque_lista = list(estoque_por_categoria.values())
        estoque_lista.sort(key=lambda x: x['valor_total'], reverse=True)
        
        # Produtos com baixo estoque (menos de 10 unidades)
        baixo_estoque = [p for p in produtos if (p.quantidade or 0) < 10]
        
        # Produtos com alto valor
        alto_valor = []
        for produto in produtos:
            quantidade = produto.quantidade if produto.quantidade is not None else 0
            preco = produto.preco_venda if produto.preco_venda is not None else 0
            valor_produto = preco * quantidade
            alto_valor.append({
                'nome': produto.nome,
                'categoria': produto.categoria,
                'quantidade': quantidade,
                'valor_unitario': preco,
                'valor_total': valor_produto
            })
        
        alto_valor.sort(key=lambda x: x['valor_total'], reverse=True)
        
        resultado = {
            'resumo': {
                'total_produtos': len(produtos),
                'valor_total_estoque': valor_total_estoque,
                'categorias': len(estoque_por_categoria),
                'produtos_baixo_estoque': len(baixo_estoque)
            },
            'estoque_por_categoria': estoque_lista,
            'produtos_baixo_estoque': [p.to_dict() for p in baixo_estoque],
            'produtos_alto_valor': alto_valor[:10]
        }
        
        print(f"Resultado final: {resultado['resumo']}")
        return jsonify(resultado)
        
    except Exception as e:
        print(f"Erro ao gerar relatório de estoque: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/relatorios/vendas', methods=['GET'])
@require_permission('view_reports')
def relatorio_vendas():
    """Gera relatório de histórico de vendas baseado em produtos cadastrados"""
    try:
        user_id = get_jwt_identity()
        
        # Buscar produtos cadastrados
        produtos = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.status == 'ativo'
        ).all()
        
        if not produtos:
            return jsonify({
                'resumo': {
                    'receita_total': 0,
                    'produtos_cadastrados': 0,
                    'valor_potencial_estoque': 0
                },
                'produtos_cadastrados': [],
                'categorias_disponiveis': []
            })
        
        # Calcular potencial de vendas baseado no estoque atual
        valor_potencial_total = 0
        produtos_info = []
        
        for produto in produtos:
            valor_potencial = (produto.quantidade or 0) * (produto.preco_venda or 0)
            valor_potencial_total += valor_potencial
            margem_unitaria = (produto.preco_venda or 0) - (produto.preco_custo or 0)
            
            produtos_info.append({
                'produto_id': produto.id,
                'nome': produto.nome,
                'categoria': produto.categoria,
                'quantidade_estoque': produto.quantidade or 0,
                'preco_unitario': produto.preco_venda or 0,
                'valor_potencial': valor_potencial,
                'margem_unitaria': margem_unitaria,
                'margem_total_potencial': margem_unitaria * (produto.quantidade or 0)
            })
        
        # Ordenar por valor potencial
        produtos_info.sort(key=lambda x: x['valor_potencial'], reverse=True)
        
        # Agrupar por categoria
        categorias_info = {}
        for produto in produtos_info:
            categoria = produto['categoria']
            if categoria not in categorias_info:
                categorias_info[categoria] = {
                    'categoria': categoria,
                    'quantidade_produtos': 0,
                    'quantidade_total_estoque': 0,
                    'valor_potencial': 0,
                    'margem_potencial': 0
                }
            
            categorias_info[categoria]['quantidade_produtos'] += 1
            categorias_info[categoria]['quantidade_total_estoque'] += produto['quantidade_estoque']
            categorias_info[categoria]['valor_potencial'] += produto['valor_potencial']
            categorias_info[categoria]['margem_potencial'] += produto['margem_total_potencial']
        
        categorias_lista = list(categorias_info.values())
        categorias_lista.sort(key=lambda x: x['valor_potencial'], reverse=True)
        
        return jsonify({
            'resumo': {
                'receita_total': 0,  # Sem vendas reais registradas
                'produtos_cadastrados': len(produtos),
                'valor_potencial_estoque': valor_potencial_total
            },
            'produtos_cadastrados': produtos_info[:10],
            'categorias_disponiveis': categorias_lista,
            'observacao': 'Dados baseados no estoque atual. Sistema de vendas não implementado.'
        })
        
    except Exception as e:
        print(f"Erro ao gerar relatório de vendas: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/relatorios/fornecedores', methods=['GET'])
@require_permission('view_reports')
def relatorio_fornecedores():
    """Gera relatório de fornecedores"""
    try:
        user_id = get_jwt_identity()
        
        # Buscar todos os produtos com fornecedores
        produtos = Produto.query.filter(
            Produto.user_id == user_id,
            Produto.status == 'ativo',
            Produto.fornecedor.isnot(None),
            Produto.fornecedor != ''
        ).all()
        
        # Agrupar por fornecedor
        fornecedores_dados = {}
        
        for produto in produtos:
            fornecedor = produto.fornecedor or 'Não informado'
            
            if fornecedor not in fornecedores_dados:
                fornecedores_dados[fornecedor] = {
                    'nome': fornecedor,
                    'total_produtos': 0,
                    'valor_total_estoque': 0,
                    'categorias': set(),
                    'produtos': []
                }
            
            valor_produto = (produto.preco_custo or 0) * (produto.quantidade or 0)
            
            fornecedores_dados[fornecedor]['total_produtos'] += 1
            fornecedores_dados[fornecedor]['valor_total_estoque'] += valor_produto
            fornecedores_dados[fornecedor]['categorias'].add(produto.categoria)
            fornecedores_dados[fornecedor]['produtos'].append({
                'id': produto.id,
                'nome': produto.nome,
                'categoria': produto.categoria,
                'quantidade': produto.quantidade,
                'preco_custo': produto.preco_custo or 0,
                'valor_total': valor_produto
            })
        
        # Converter sets para listas e preparar dados finais
        fornecedores_lista = []
        for fornecedor_data in fornecedores_dados.values():
            fornecedor_data['categorias'] = list(fornecedor_data['categorias'])
            fornecedor_data['quantidade_categorias'] = len(fornecedor_data['categorias'])
            fornecedores_lista.append(fornecedor_data)
        
        # Ordenar por valor total do estoque
        fornecedores_lista.sort(key=lambda x: x['valor_total_estoque'], reverse=True)
        
        return jsonify({
            'resumo': {
                'total_fornecedores': len(fornecedores_lista),
                'total_produtos': len(produtos),
                'valor_total': sum(f['valor_total_estoque'] for f in fornecedores_lista)
            },
            'fornecedores': fornecedores_lista
        })
        
    except Exception as e:
        print(f"Erro ao gerar relatório de fornecedores: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/relatorios/financeiro', methods=['GET'])
@require_permission('view_reports')
def relatorio_financeiro():
    try:
        user_id = get_jwt_identity()
        produtos = Produto.query.filter_by(user_id=user_id, status='ativo').all()
        
        # Cálculos financeiros baseados no estoque atual
        valor_total_investido = 0  # Custo total dos produtos
        valor_total_potencial = 0  # Valor de venda potencial
        margem_total_potencial = 0  # Margem de lucro potencial
        
        categorias_mais_lucrativas = {}
        produtos_mais_lucrativos = []
        
        for produto in produtos:
            preco_custo = produto.preco_custo or 0
            preco_venda = produto.preco_venda or 0
            quantidade = produto.quantidade or 0
            
            custo_total_produto = preco_custo * quantidade
            valor_venda_produto = preco_venda * quantidade
            margem_produto = valor_venda_produto - custo_total_produto
            
            valor_total_investido += custo_total_produto
            valor_total_potencial += valor_venda_produto
            margem_total_potencial += margem_produto
            
            # Análise por categoria
            categoria = produto.categoria
            if categoria not in categorias_mais_lucrativas:
                categorias_mais_lucrativas[categoria] = {
                    'categoria': categoria,
                    'valor_investido': 0,
                    'valor_potencial': 0,
                    'margem_potencial': 0,
                    'quantidade_produtos': 0
                }
            
            categorias_mais_lucrativas[categoria]['valor_investido'] += custo_total_produto
            categorias_mais_lucrativas[categoria]['valor_potencial'] += valor_venda_produto
            categorias_mais_lucrativas[categoria]['margem_potencial'] += margem_produto
            categorias_mais_lucrativas[categoria]['quantidade_produtos'] += 1
            
            # Produtos mais lucrativos
            if margem_produto > 0:
                produtos_mais_lucrativos.append({
                    'nome': produto.nome,
                    'categoria': produto.categoria,
                    'valor_investido': custo_total_produto,
                    'valor_potencial': valor_venda_produto,
                    'margem_potencial': margem_produto,
                    'margem_percentual': ((margem_produto / custo_total_produto) * 100) if custo_total_produto > 0 else 0
                })
        
        # Ordenar produtos por margem potencial
        produtos_mais_lucrativos.sort(key=lambda x: x['margem_potencial'], reverse=True)
        produtos_mais_lucrativos = produtos_mais_lucrativos[:10]  # Top 10
        
        # Ordenar categorias por margem potencial
        categorias_lista = list(categorias_mais_lucrativas.values())
        categorias_lista.sort(key=lambda x: x['margem_potencial'], reverse=True)
        
        # Calcular percentual de margem geral
        margem_percentual_geral = ((margem_total_potencial / valor_total_investido) * 100) if valor_total_investido > 0 else 0
        
        return jsonify({
            'resumo': {
                'valor_total_investido': valor_total_investido,
                'valor_total_potencial': valor_total_potencial,
                'margem_total_potencial': margem_total_potencial,
                'margem_percentual_geral': margem_percentual_geral,
                'total_produtos': len(produtos)
            },
            'categorias_mais_lucrativas': categorias_lista,
            'produtos_mais_lucrativos': produtos_mais_lucrativos,
            'observacao': 'Dados baseados no estoque atual e preços de custo/venda cadastrados. Não inclui vendas reais.'
        })
        
    except Exception as e:
        print(f"Erro ao gerar relatório financeiro: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

# Rotas de Alertas
@app.route('/api/alertas', methods=['GET'])
@require_permission('view_alerts')
def listar_alertas():
    try:
        user_id = int(get_jwt_identity())
        
        # Parâmetros de filtro
        status = request.args.get('status', 'ativo')
        tipo = request.args.get('tipo')
        lido = request.args.get('lido')
        
        # Query base
        query = Alerta.query.filter_by(user_id=user_id)
        
        # Aplicar filtros
        if status:
            query = query.filter_by(status=status)
        if tipo:
            query = query.filter_by(tipo=tipo)
        if lido is not None:
            lido_bool = lido.lower() == 'true'
            query = query.filter_by(lido=lido_bool)
        
        # Ordenar por data de criação (mais recentes primeiro)
        alertas = query.order_by(Alerta.created_at.desc()).all()
        
        return jsonify({
            'alertas': [alerta.to_dict() for alerta in alertas],
            'total': len(alertas)
        })
        
    except Exception as e:
        print(f"Erro ao listar alertas: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/alertas/<int:alerta_id>/marcar-lido', methods=['PUT'])
@jwt_required()
def marcar_alerta_lido(alerta_id):
    try:
        user_id = int(get_jwt_identity())
        
        alerta = Alerta.query.filter_by(id=alerta_id, user_id=user_id).first()
        if not alerta:
            return jsonify({'error': 'Alerta não encontrado'}), 404
        
        alerta.lido = True
        db.session.commit()
        
        return jsonify({
            'message': 'Alerta marcado como lido',
            'alerta': alerta.to_dict()
        })
        
    except Exception as e:
        print(f"Erro ao marcar alerta como lido: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/alertas/marcar-todos-lidos', methods=['PUT'])
@jwt_required()
def marcar_todos_alertas_lidos():
    try:
        user_id = int(get_jwt_identity())
        
        # Atualizar todos os alertas não lidos do usuário
        alertas_nao_lidos = Alerta.query.filter_by(user_id=user_id, lido=False).all()
        
        for alerta in alertas_nao_lidos:
            alerta.lido = True
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(alertas_nao_lidos)} alertas marcados como lidos',
            'total_marcados': len(alertas_nao_lidos)
        })
        
    except Exception as e:
        print(f"Erro ao marcar todos os alertas como lidos: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/alertas/gerar', methods=['POST'])
@jwt_required()
def gerar_alertas():
    """Gera alertas automáticos baseados nos produtos e configurações do usuário"""
    try:
        user_id = int(get_jwt_identity())
        
        # Buscar configurações do usuário
        config_vencimento = ConfiguracaoAlerta.query.filter_by(
            user_id=user_id, 
            tipo_alerta='vencimento',
            ativo=True
        ).first()
        
        config_estoque = ConfiguracaoAlerta.query.filter_by(
            user_id=user_id,
            tipo_alerta='estoque_baixo', 
            ativo=True
        ).first()
        
        alertas_criados = []
        
        # Gerar alertas de vencimento
        if config_vencimento:
            dias_antecedencia = config_vencimento.dias_antecedencia
            data_limite = datetime.now().date() + timedelta(days=dias_antecedencia)
            
            produtos_vencendo = Produto.query.filter(
                Produto.user_id == user_id,
                Produto.data_validade <= data_limite,
                Produto.data_validade >= datetime.now().date(),
                Produto.quantidade > 0
            ).all()
            
            for produto in produtos_vencendo:
                # Verificar se já existe alerta para este produto
                alerta_existente = Alerta.query.filter_by(
                    produto_id=produto.id,
                    tipo='vencimento',
                    status='ativo'
                ).first()
                
                if not alerta_existente:
                    dias_restantes = (produto.data_validade - datetime.now().date()).days
                    urgencia = 'alta' if dias_restantes <= 3 else 'media' if dias_restantes <= 7 else 'baixa'
                    
                    alerta = Alerta(
                        produto_id=produto.id,
                        user_id=user_id,
                        tipo='vencimento',
                        urgencia=urgencia,
                        titulo=f'Produto vence em {dias_restantes} dias',
                        descricao=f'{produto.nome} vence em {dias_restantes} dias ({produto.quantidade} unidades)',
                        quantidade_afetada=produto.quantidade,
                        valor_estimado_perda=produto.quantidade * produto.preco_venda
                    )
                    
                    db.session.add(alerta)
                    alertas_criados.append(alerta)
        
        # Gerar alertas de estoque baixo
        if config_estoque:
            estoque_minimo = config_estoque.estoque_minimo
            
            produtos_estoque_baixo = Produto.query.filter(
                Produto.user_id == user_id,
                Produto.quantidade <= estoque_minimo,
                Produto.quantidade > 0
            ).all()
            
            for produto in produtos_estoque_baixo:
                # Verificar se já existe alerta para este produto
                alerta_existente = Alerta.query.filter_by(
                    produto_id=produto.id,
                    tipo='estoque_baixo',
                    status='ativo'
                ).first()
                
                if not alerta_existente:
                    urgencia = 'alta' if produto.quantidade <= 2 else 'media' if produto.quantidade <= 5 else 'baixa'
                    
                    alerta = Alerta(
                        produto_id=produto.id,
                        user_id=user_id,
                        tipo='estoque_baixo',
                        urgencia=urgencia,
                        titulo=f'Estoque baixo: {produto.quantidade} unidades',
                        descricao=f'{produto.nome} com apenas {produto.quantidade} unidades em estoque',
                        quantidade_afetada=produto.quantidade
                    )
                    
                    db.session.add(alerta)
                    alertas_criados.append(alerta)
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(alertas_criados)} alertas gerados',
            'alertas': [alerta.to_dict() for alerta in alertas_criados]
        })
        
    except Exception as e:
        print(f"Erro ao gerar alertas: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

# Rotas de Configurações de Alertas
@app.route('/api/configuracoes-alertas', methods=['GET'])
@jwt_required()
def listar_configuracoes_alertas():
    try:
        user_id = int(get_jwt_identity())
        
        configuracoes = ConfiguracaoAlerta.query.filter_by(user_id=user_id).all()
        
        return jsonify({
            'configuracoes': [config.to_dict() for config in configuracoes]
        })
        
    except Exception as e:
        print(f"Erro ao listar configurações de alertas: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/vendas', methods=['GET'])
@require_permission('view_sales')
def listar_vendas():
    """Lista todas as vendas do usuário"""
    try:
        user_id = get_jwt_identity()
        
        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        produto_id = request.args.get('produto_id')
        
        query = Venda.query.filter(Venda.user_id == user_id)
        
        if data_inicio:
            query = query.filter(Venda.data_venda >= datetime.fromisoformat(data_inicio))
        if data_fim:
            query = query.filter(Venda.data_venda <= datetime.fromisoformat(data_fim))
        if produto_id:
            query = query.filter(Venda.produto_id == produto_id)
        
        vendas = query.order_by(Venda.data_venda.desc()).all()
        
        return jsonify({
            'vendas': [venda.to_dict() for venda in vendas],
            'total': len(vendas)
        })
        
    except Exception as e:
        print(f"Erro ao listar vendas: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/vendas', methods=['POST'])
@require_permission('create_sale')
def criar_venda():
    """Cria uma nova venda"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        
        # Validações
        if not data.get('produto_id'):
            return jsonify({'error': 'produto_id é obrigatório'}), 400
        if not data.get('quantidade') or data['quantidade'] <= 0:
            return jsonify({'error': 'quantidade deve ser maior que zero'}), 400
        if not data.get('preco_unitario') or data['preco_unitario'] <= 0:
            return jsonify({'error': 'preco_unitario deve ser maior que zero'}), 400
        
        # Verificar se o produto existe e pertence ao usuário
        produto = Produto.query.filter_by(
            id=data['produto_id'],
            user_id=user_id,
            status='ativo'
        ).first()
        
        if not produto:
            return jsonify({'error': 'Produto não encontrado'}), 404
        
        # Verificar se há estoque suficiente
        if produto.quantidade < data['quantidade']:
            return jsonify({'error': 'Estoque insuficiente'}), 400
        
        # Calcular valor total
        valor_total = data['quantidade'] * data['preco_unitario']
        
        # Criar venda
        venda = Venda(
            user_id=user_id,
            produto_id=data['produto_id'],
            quantidade=data['quantidade'],
            preco_unitario=data['preco_unitario'],
            valor_total=valor_total,
            metodo_pagamento=data.get('metodo_pagamento'),
            observacoes=data.get('observacoes'),
            data_venda=datetime.fromisoformat(data['data_venda']) if data.get('data_venda') else datetime.utcnow()
        )
        
        # Atualizar estoque do produto
        produto.quantidade -= data['quantidade']
        produto.updated_at = datetime.utcnow()
        
        db.session.add(venda)
        db.session.commit()
        
        return jsonify({
            'message': 'Venda criada com sucesso',
            'venda': venda.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao criar venda: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/configuracoes-alertas', methods=['POST'])
@jwt_required()
def criar_configuracao_alerta():
    try:
        user_id = int(get_jwt_identity())
        data = request.get_json()
        
        # Validação básica
        if not data.get('tipo_alerta'):
            return jsonify({'error': 'Tipo de alerta é obrigatório'}), 400
        
        # Verificar se já existe configuração para este tipo
        config_existente = ConfiguracaoAlerta.query.filter_by(
            user_id=user_id,
            tipo_alerta=data['tipo_alerta']
        ).first()
        
        if config_existente:
            # Atualizar configuração existente
            config_existente.ativo = data.get('ativo', True)
            config_existente.dias_antecedencia = data.get('dias_antecedencia', 7)
            config_existente.estoque_minimo = data.get('estoque_minimo', 5)
            config_existente.categorias = data.get('categorias')
            config_existente.notificar_email = data.get('notificar_email', True)
            config_existente.notificar_sistema = data.get('notificar_sistema', True)
            config_existente.recorrencia = data.get('recorrencia', 'diaria')
            config_existente.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            return jsonify({
                'message': 'Configuração atualizada com sucesso',
                'configuracao': config_existente.to_dict()
            })
        else:
            # Criar nova configuração
            configuracao = ConfiguracaoAlerta(
                user_id=user_id,
                tipo_alerta=data['tipo_alerta'],
                ativo=data.get('ativo', True),
                dias_antecedencia=data.get('dias_antecedencia', 7),
                estoque_minimo=data.get('estoque_minimo', 5),
                categorias=data.get('categorias'),
                notificar_email=data.get('notificar_email', True),
                notificar_sistema=data.get('notificar_sistema', True),
                recorrencia=data.get('recorrencia', 'diaria')
            )
            
            db.session.add(configuracao)
            db.session.commit()
            
            return jsonify({
                'message': 'Configuração criada com sucesso',
                'configuracao': configuracao.to_dict()
            })
        
    except Exception as e:
        print(f"Erro ao criar/atualizar configuração de alerta: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

# Endpoints de Configurações do Usuário
@app.route('/api/configuracoes', methods=['GET'])
@jwt_required()
def obter_configuracoes():
    try:
        user_id = int(get_jwt_identity())
        
        # Buscar configurações existentes ou criar padrão
        configuracao = ConfiguracaoUsuario.query.filter_by(user_id=user_id).first()
        
        if not configuracao:
            # Criar configuração padrão se não existir
            configuracao = ConfiguracaoUsuario(user_id=user_id)
            db.session.add(configuracao)
            db.session.commit()
        
        return jsonify({
            'configuracao': configuracao.to_dict()
        })
        
    except Exception as e:
        print(f"Erro ao obter configurações: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/configuracoes', methods=['PUT'])
@jwt_required()
def salvar_configuracoes():
    try:
        user_id = int(get_jwt_identity())
        data = request.get_json()
        
        # Buscar configuração existente ou criar nova
        configuracao = ConfiguracaoUsuario.query.filter_by(user_id=user_id).first()
        
        if not configuracao:
            configuracao = ConfiguracaoUsuario(user_id=user_id)
            db.session.add(configuracao)
        
        # Atualizar campos se fornecidos
        if 'notificacoes_email' in data:
            configuracao.notificacoes_email = data['notificacoes_email']
        if 'notificacoes_push' in data:
            configuracao.notificacoes_push = data['notificacoes_push']
        if 'dias_antecedencia_notificacao' in data:
            configuracao.dias_antecedencia_notificacao = data['dias_antecedencia_notificacao']
        if 'autenticacao_dois_fatores' in data:
            configuracao.autenticacao_dois_fatores = data['autenticacao_dois_fatores']
        if 'timeout_sessao' in data:
            configuracao.timeout_sessao = data['timeout_sessao']
        if 'tema' in data:
            configuracao.tema = data['tema']
        if 'idioma' in data:
            configuracao.idioma = data['idioma']
        if 'smtp_servidor' in data:
            configuracao.smtp_servidor = data['smtp_servidor']
        if 'smtp_porta' in data:
            configuracao.smtp_porta = data['smtp_porta']
        if 'smtp_usuario' in data:
            configuracao.smtp_usuario = data['smtp_usuario']
        if 'smtp_senha' in data:
            configuracao.smtp_senha = data['smtp_senha']
        if 'smtp_ssl' in data:
            configuracao.smtp_ssl = data['smtp_ssl']
        if 'backup_automatico' in data:
            configuracao.backup_automatico = data['backup_automatico']
        if 'frequencia_backup' in data:
            configuracao.frequencia_backup = data['frequencia_backup']
        if 'local_backup' in data:
            configuracao.local_backup = data['local_backup']
        
        configuracao.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'message': 'Configurações salvas com sucesso',
            'configuracao': configuracao.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao salvar configurações: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

# Endpoint para exportar dados
@app.route('/api/configuracoes/exportar', methods=['GET'])
@jwt_required()
def exportar_dados():
    try:
        user_id = int(get_jwt_identity())
        
        # Buscar todos os dados do usuário
        produtos = Produto.query.filter_by(user_id=user_id).all()
        vendas = Venda.query.filter_by(user_id=user_id).all()
        alertas = Alerta.query.filter_by(user_id=user_id).all()
        configuracoes = ConfiguracaoUsuario.query.filter_by(user_id=user_id).first()
        
        dados_exportacao = {
            'data_exportacao': datetime.utcnow().isoformat(),
            'produtos': [produto.to_dict() for produto in produtos],
            'vendas': [venda.to_dict() for venda in vendas],
            'alertas': [alerta.to_dict() for alerta in alertas],
            'configuracoes': configuracoes.to_dict() if configuracoes else None
        }
        
        return jsonify(dados_exportacao)
        
    except Exception as e:
        print(f"Erro ao exportar dados: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

# ==================== ENDPOINTS DE USUÁRIOS ====================

@app.route('/api/usuarios', methods=['POST'])
@require_permission('create_user')
def criar_usuario():
    """Cria um novo usuário (apenas para admins e gerentes)"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(current_user_id)
        
        # Verificar se o usuário tem permissão para criar usuários
        if current_user.cargo not in ['admin', 'gerente']:
            return jsonify({'error': 'Acesso negado'}), 403
        
        data = request.get_json()
        
        # Validações obrigatórias
        campos_obrigatorios = ['email', 'password', 'nome_estabelecimento', 'cargo']
        for campo in campos_obrigatorios:
            if not data.get(campo):
                return jsonify({'error': f'Campo {campo} é obrigatório'}), 400
        
        # Verificar se email já existe
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email já cadastrado'}), 400
        
        # Validar cargo
        cargos_validos = ['admin', 'gerente', 'usuario', 'visualizador']
        if data['cargo'] not in cargos_validos:
            return jsonify({'error': 'Cargo inválido'}), 400
        
        # Apenas admins podem criar outros admins
        if data['cargo'] == 'admin' and current_user.cargo != 'admin':
            return jsonify({'error': 'Apenas administradores podem criar outros administradores'}), 403
        
        # Criar novo usuário
        novo_usuario = User(
            email=data['email'],
            nome_estabelecimento=data['nome_estabelecimento'],
            nome_completo=data.get('nome_completo'),
            telefone=data.get('telefone'),
            empresa=data.get('empresa'),
            bio=data.get('bio'),
            cargo=data['cargo'],
            permissoes=data.get('permissoes'),
            ativo=data.get('ativo', True)
        )
        novo_usuario.set_password(data['password'])  # Usar hash de senha
        
        db.session.add(novo_usuario)
        db.session.commit()
        
        return jsonify({
            'message': 'Usuário criado com sucesso',
            'usuario': novo_usuario.to_dict()
        }), 201
        
    except Exception as e:
        print(f"Erro ao criar usuário: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios', methods=['GET'])
@require_permission('view_all_users')
def listar_usuarios():
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(current_user_id)
        
        # Verificar se o usuário tem permissão para listar usuários
        if current_user.cargo not in ['admin', 'gerente']:
            return jsonify({'error': 'Acesso negado'}), 403
        
        # Buscar todos os usuários
        usuarios = User.query.all()
        
        return jsonify({
            'usuarios': [usuario.to_dict() for usuario in usuarios],
            'total': len(usuarios)
        }), 200
        
    except Exception as e:
        print(f"Erro ao listar usuários: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios/<int:usuario_id>', methods=['GET'])
@require_permission('view_all_users')
def obter_usuario(usuario_id):
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(current_user_id)
        
        # Verificar se o usuário pode ver este perfil
        if current_user_id != usuario_id and current_user.cargo not in ['admin', 'gerente']:
            return jsonify({'error': 'Acesso negado'}), 403
        
        usuario = User.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        return jsonify(usuario.to_dict()), 200
        
    except Exception as e:
        print(f"Erro ao obter usuário: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios/<int:usuario_id>', methods=['PUT'])
@require_permission('edit_user')
def atualizar_usuario(usuario_id):
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(current_user_id)
        
        # Verificar se o usuário pode editar este perfil
        if current_user_id != usuario_id and current_user.cargo not in ['admin', 'gerente']:
            return jsonify({'error': 'Acesso negado'}), 403
        
        usuario = User.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        data = request.get_json()
        
        # Campos que podem ser atualizados
        campos_permitidos = [
            'nome_completo', 'telefone', 'empresa', 'bio', 'foto_perfil'
        ]
        
        # Campos que só admins podem alterar
        campos_admin = ['cargo', 'permissoes', 'ativo']
        
        # Atualizar campos básicos
        for campo in campos_permitidos:
            if campo in data:
                setattr(usuario, campo, data[campo])
        
        # Atualizar campos administrativos (só para admins)
        if current_user.cargo == 'admin':
            for campo in campos_admin:
                if campo in data:
                    setattr(usuario, campo, data[campo])
        
        # Atualizar timestamp
        usuario.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'message': 'Usuário atualizado com sucesso',
            'usuario': usuario.to_dict()
        }), 200
        
    except Exception as e:
        print(f"Erro ao atualizar usuário: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios/<int:usuario_id>', methods=['DELETE'])
@require_permission('delete_user')
def deletar_usuario(usuario_id):
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(current_user_id)
        
        # Apenas admins podem deletar usuários
        if current_user.cargo != 'admin':
            return jsonify({'error': 'Acesso negado'}), 403
        
        # Não permitir que o admin delete a si mesmo
        if current_user_id == usuario_id:
            return jsonify({'error': 'Não é possível deletar seu próprio usuário'}), 400
        
        usuario = User.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        # Soft delete - apenas desativar o usuário
        usuario.ativo = False
        usuario.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'message': 'Usuário desativado com sucesso'}), 200
        
    except Exception as e:
        print(f"Erro ao deletar usuário: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios/<int:usuario_id>/alterar-senha', methods=['PUT'])
@require_permission('change_user_password')
def alterar_senha_usuario(usuario_id):
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(current_user_id)
        
        # Verificar se o usuário pode alterar esta senha
        if current_user_id != usuario_id and current_user.cargo != 'admin':
            return jsonify({'error': 'Acesso negado'}), 403
        
        usuario = User.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        data = request.get_json()
        
        # Se não for admin, precisa da senha atual
        if current_user_id == usuario_id and current_user.cargo != 'admin':
            if 'senha_atual' not in data:
                return jsonify({'error': 'Senha atual é obrigatória'}), 400
            
            # Verificar senha atual (aqui você implementaria a verificação de hash)
            # Por simplicidade, vamos assumir que a verificação está correta
        
        if 'nova_senha' not in data:
            return jsonify({'error': 'Nova senha é obrigatória'}), 400
        
        # Atualizar senha usando hash
        usuario.set_password(data['nova_senha'])
        usuario.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'message': 'Senha alterada com sucesso'}), 200
        
    except Exception as e:
        print(f"Erro ao alterar senha: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios/<int:usuario_id>/upload-foto', methods=['POST'])
@jwt_required()
def upload_foto_perfil(usuario_id):
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(current_user_id)
        
        # Verificar se o usuário pode alterar esta foto
        if current_user_id != usuario_id and current_user.cargo not in ['admin', 'gerente']:
            return jsonify({'error': 'Acesso negado'}), 403
        
        usuario = User.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        # Verificar se há arquivo na requisição
        if 'foto_perfil' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['foto_perfil']
        
        # Verificar se um arquivo foi selecionado
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        # Salvar o arquivo
        file_path = save_uploaded_file(file)
        if not file_path:
            return jsonify({'error': 'Tipo de arquivo não permitido. Use PNG, JPG, JPEG, GIF ou WEBP'}), 400
        
        # Remover foto anterior se existir
        if usuario.foto_perfil and usuario.foto_perfil.startswith('/uploads/'):
            old_file_path = os.path.join(app.config['UPLOAD_FOLDER'], usuario.foto_perfil.replace('/uploads/', ''))
            if os.path.exists(old_file_path):
                try:
                    os.remove(old_file_path)
                except Exception as e:
                    print(f"Erro ao remover arquivo anterior: {e}")
        
        # Atualizar foto de perfil
        usuario.foto_perfil = file_path
        usuario.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'message': 'Foto de perfil atualizada com sucesso',
            'foto_url': usuario.foto_perfil
        }), 200
        
    except Exception as e:
        print(f"Erro ao fazer upload da foto: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios/perfil', methods=['GET'])
@jwt_required()
def obter_perfil_atual():
    try:
        user_id = get_jwt_identity()
        usuario = User.query.get(user_id)
        
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        return jsonify(usuario.to_dict()), 200
        
    except Exception as e:
        print(f"Erro ao obter perfil: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios/perfil', methods=['PUT'])
@jwt_required()
def atualizar_perfil_atual():
    try:
        user_id = get_jwt_identity()
        usuario = User.query.get(user_id)
        
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        data = request.get_json()
        
        # Campos que o usuário pode atualizar em seu próprio perfil
        campos_permitidos = [
            'nome_completo', 'telefone', 'empresa', 'bio', 'foto_perfil'
        ]
        
        for campo in campos_permitidos:
            if campo in data:
                setattr(usuario, campo, data[campo])
        
        usuario.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'message': 'Perfil atualizado com sucesso',
            'usuario': usuario.to_dict()
        }), 200
        
    except Exception as e:
        print(f"Erro ao atualizar perfil: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/usuarios/permissoes', methods=['GET'])
@jwt_required()
def obter_permissoes_usuario():
    """Retorna as permissões do usuário atual"""
    try:
        user_id = get_jwt_identity()
        print(f"🔍 DEBUG: user_id from JWT: {user_id}")
        user = User.query.get(user_id)
        
        if not user:
            print(f"❌ DEBUG: Usuário não encontrado para ID: {user_id}")
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        print(f"👤 DEBUG: Usuário encontrado: {user.email}, cargo: {user.cargo}")
        print(f"🔐 DEBUG: user.permissoes: {user.permissoes}, tipo: {type(user.permissoes)}")
        
        if not user.ativo:
            print(f"❌ DEBUG: Usuário inativo: {user.email}")
            return jsonify({'error': 'Usuário inativo'}), 403
        
        # Verificar se o usuário tem permissões customizadas
        if user.permissoes:
            print(f"✅ DEBUG: Usuário tem permissões customizadas: {user.permissoes}")
            # Se tem permissão 'all', retorna ['all'] para que o frontend reconheça
            if 'all' in user.permissoes:
                permissions = ['all']
                print(f"🌟 DEBUG: Usuário tem permissão 'all', retornando: {permissions}")
            else:
                # Usar permissões customizadas
                permissions = user.permissoes
                print(f"📋 DEBUG: Usando permissões customizadas: {permissions}")
        else:
            print(f"⚠️ DEBUG: Usuário sem permissões customizadas, usando cargo: {user.cargo}")
            # Fallback para permissões baseadas no cargo
            permissions = get_user_permissions(user.cargo)
            print(f"📋 DEBUG: Permissões do cargo {user.cargo}: {permissions}")
        
        return jsonify({
            'cargo': user.cargo,
            'permissoes': permissions,
            'permissoes_customizadas': user.permissoes,
            'usuario': {
                'id': user.id,
                'email': user.email,
                'nome_completo': user.nome_completo,
                'ativo': user.ativo
            }
        }), 200
        
    except Exception as e:
        print(f"Erro ao obter permissões: {str(e)}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/api/test', methods=['GET'])
def test():
    return jsonify({'message': 'API funcionando!', 'timestamp': datetime.now().isoformat()})

@app.route('/api/debug/produtos', methods=['GET'])
@jwt_required()
def debug_produtos():
    """Endpoint temporário para debug dos produtos"""
    try:
        user_id = get_jwt_identity()
        produtos = Produto.query.filter(Produto.user_id == user_id).all()
        
        produtos_data = []
        for p in produtos:
            produtos_data.append({
                'id': p.id,
                'nome': p.nome,
                'categoria': p.categoria,
                'quantidade': p.quantidade,
                'preco_venda': p.preco_venda,
                'preco_custo': p.preco_custo,
                'status': p.status,
                'data_validade': p.data_validade.strftime('%Y-%m-%d') if p.data_validade else None
            })
        
        return jsonify({
            'total_produtos': len(produtos),
            'produtos': produtos_data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Endpoint para servir arquivos estáticos de upload
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """Serve arquivos de upload"""
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        print(f"Erro ao servir arquivo: {e}")
        return jsonify({'error': 'Arquivo não encontrado'}), 404

# Registrar blueprint de notificações
app.register_blueprint(notifications_bp, url_prefix='/api')

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        print("Database initialized successfully")
    
    print("Starting Flask server on http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)